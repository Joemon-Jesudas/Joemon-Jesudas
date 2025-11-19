PA_VALIDATOR_TEMPLATE = """
You are an Allianz Contract Validation Expert.

You must validate an Allianz Project Agreement (PA) based strictly on the following rules:

==========================================
📌 PA VALIDATION RULES
==========================================

1. TEMPLATE CLASSIFICATION
- Identify contract category as IT, Non-IT, or Marketing.
- IT template header must contain: “Agreement for IT Projects and Services”
- Non-IT keywords: Consulting Agreement, Consulting, teaching, advising, coaching
- Marketing keywords: Marketing, Market Insight, Media  
Output: value + Correct/Missing/Mismatch

2. NAME AND ADDRESS VALIDATION
- Allianz address must be either:
  “Allianz SE, Königinstrasse 28, 80802 München Germany”
  OR
  “Allianz Technology SE, Königinstrasse 28, 80802 München Germany”
- Supplier name & address must be extracted fully  
Output: value + Correct/Missing/Mismatch

3. POINTS OF CONTACT VALIDATION
Customer Contact:
- Extract: surname, first name, phone, email
- Email must contain “Allianz” → otherwise Mismatch

Contractor Contact:
- Email must contain supplier domain → otherwise Mismatch

4. PLACE OF PERFORMANCE
- If “Others” is selected → details must be present
- If default → return Correct  
Output: Correct/Missing

5. SUBCONTRACTOR DETAILS  
Always extract the list. If none exists → Missing

6. REMUNERATION
- Identify which checkbox is selected (Fixed price / Based on time / Based on rate card)
- Extract EUR amounts & table content
- If option requires table and table is missing → Missing

7. INVOICING
- Fixed price → only options 2 or 3 allowed → otherwise Mismatch
- Time-based remuneration → any option allowed

8. VAT
- If supplier is German → “local contractor” must be selected
- If supplier is NOT German → “reverse charge” must be selected  
Output: Correct/Missing/Mismatch

9. INVOICE ADDRESS
- Must match customer OE (AZSE / AZTECH)
- Or match: “Dieselstraße 8 85774 Unterföhring Germany”

10. DATA PROTECTION / INFO SECURITY / OUTSOURCING
- If checkbox marked “Yes” → attachment must exist → Correct
- If attachment missing → Missing

11. TERMS
- Extract Start Date & End Date
- These fields MUST exist → otherwise Missing

12. SIGNATURE VALIDATION
- Allianz SE buyer → 3 signatures (2 Allianz + 1 supplier)
- Multi-year contract → requires 3 signatures
- Vendor consolidation → requires 4 signatures  
Output: count + Correct/Missing/Mismatch

13. STRIKETHROUGH CHECK
- The contract MUST NOT contain strikethrough text  
Output: Correct/Missing/Mismatch

==========================================
📌 OUTPUT SPECIFICATION
==========================================

ALWAYS return the result in **THIS EXACT JSON FORMAT**:

{
  "Template Category": {"value": "", "status": "", "evidence": ""},
  "Allianz Address": {"value": "", "status": "", "evidence": ""},
  "Supplier Address": {"value": "", "status": "", "evidence": ""},
  "Customer Contact": {"value": "", "status": "", "evidence": ""},
  "Contractor Contact": {"value": "", "status": "", "evidence": ""},
  "Place of Performance": {"value": "", "status": "", "evidence": ""},
  "Subcontractors": {"value": "", "status": "", "evidence": ""},
  "Remuneration": {"value": "", "status": "", "evidence": ""},
  "Invoicing": {"value": "", "status": "", "evidence": ""},
  "VAT": {"value": "", "status": "", "evidence": ""},
  "Invoice Address": {"value": "", "status": "", "evidence": ""},
  "Data Protection": {"value": "", "status": "", "evidence": ""},
  "Information Security": {"value": "", "status": "", "evidence": ""},
  "Outsourcing": {"value": "", "status": "", "evidence": ""},
  "Start Date": {"value": "", "status": "", "evidence": ""},
  "End Date": {"value": "", "status": "", "evidence": ""},
  "Termination for Convenience": {"value": "", "status": "", "evidence": ""},
  "Attachment 1: Service Description": {"value": "", "status": "", "evidence": ""},
  "Signatures": {"value": "", "status": "", "evidence": ""},
  "Strikethrough Check": {"value": "", "status": "", "evidence": ""}
}

RULES:
- ALWAYS follow JSON format exactly
- NEVER hallucinate values not in contract
- ALWAYS provide evidence from the contract
- ALWAYS perform strict rule-based evaluations
"""


pip install openpyxl

#validation_to_excel.py
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def status_color(status):
    if status == "Correct":
        return "C6EFCE"   # Green
    if status == "Missing":
        return "FFEB9C"   # Yellow
    if status == "Mismatch":
        return "FFC7CE"   # Red
    return "FFFFFF"

def convert_validation_to_excel(json_data, output_path="validation_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Results"

    # Header
    headers = ["Validation Item", "Extracted Value", "Status", "Evidence"]
    ws.append(headers)

    # Bold header
    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Populate rows
    for section, result in json_data.items():
        value = result.get("value", "")
        status = result.get("status", "")
        evidence = result.get("evidence", "")

        row = [section, value, status, evidence]
        ws.append(row)

        # Colour coding based on status
        fill = PatternFill(start_color=status_color(status), fill_type="solid")
        ws.cell(row=ws.max_row, column=3).fill = fill

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_path)
    print(f"Excel report saved: {output_path}")


#STEP 2 — Convert JSON → PDF Report|
pip install reportlab

#validation_to_pdf.py
import json
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

def convert_validation_to_pdf(json_data, output_path="validation_report.pdf"):
    doc = SimpleDocTemplate(output_path, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Table Header
    data = [["Validation Item", "Extracted Value", "Status", "Evidence"]]

    # Table Rows
    for section, result in json_data.items():
        data.append([
            section,
            result.get("value", ""),
            result.get("status", ""),
            result.get("evidence", "")
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 11),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))

    # Color-code status column
    for i in range(1, len(data)):
        status = data[i][2]
        color = colors.white
        if status == "Correct":
            color = colors.lightgreen
        elif status == "Missing":
            color = colors.lightyellow
        elif status == "Mismatch":
            color = colors.pink

        table.setStyle(TableStyle([
            ("BACKGROUND", (2, i), (2, i), color)
        ]))

    elements.append(table)
    doc.build(elements)
    print(f"PDF report saved: {output_path}")
#STEP 3 — Use Both Functions
import json

# Load LLM output JSON (from GPT-4o)
with open("template.json", "r") as f:
    validation_json = json.load(f)

# Generate Excel
convert_validation_to_excel(validation_json, "PA_Validation_Report.xlsx")

# Generate PDF
convert_validation_to_pdf(validation_json, "PA_Validation_Report.pdf")
